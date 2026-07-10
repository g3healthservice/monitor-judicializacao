"""Relatorio XLSX por municipio. Formatacao sobria (apreciacao por TCE/TCU)."""
from __future__ import annotations

from pathlib import Path
from typing import Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..config.constants import (
    FAIXA_ESTADUAL,
    FAIXA_FEDERAL,
    FAIXA_LOCAL,
    FAIXA_ONCOLOGICO,
)

_AZUL = "1F3864"
_CINZA = "D9D9D9"
_BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)

_FAIXA_LABEL = {
    FAIXA_FEDERAL: "Federal - Uniao 100%",
    FAIXA_ESTADUAL: "Estadual - Uniao 65%",
    FAIXA_ONCOLOGICO: "Oncologico - Uniao 80%",
    FAIXA_LOCAL: "Local - sem ressarcimento",
}


def _hdr(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=_AZUL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDA


def _brl(v: Optional[float]) -> str:
    if v is None:
        return "-"
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def gerar_xlsx(metricas: Dict, destino: Path, municipio_nome: str = "") -> Path:
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Aba Resumo ---
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Monitoramento de Judicializacao de Medicamentos"
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)
    ws["A2"] = f"Municipio: {municipio_nome or metricas['municipio_ibge']} ({metricas.get('uf') or '-'})"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = "Enquadramento: Tema 1.234/STF (SV 60 e 61)"
    ws["A3"].font = Font(italic=True, size=9, color="808080")

    linha = 5
    resumo = [
        ("Comarca", metricas.get("comarca") or "-"),
        ("Numero de processos", metricas["n_processos"]),
        ("Valor total (causa)", _brl(metricas["valor_total_causa"])),
        ("Dinheiro na mesa (ressarcivel pela Uniao)", _brl(metricas["valor_total_ressarcivel"])),
        ("Tempo medio de tramitacao (dias)", metricas["tempo_medio_tramitacao_dias"] or "-"),
    ]
    for rotulo, valor in resumo:
        ws.cell(row=linha, column=1, value=rotulo).font = Font(bold=True)
        c = ws.cell(row=linha, column=2, value=valor)
        c.alignment = Alignment(horizontal="right")
        if "Dinheiro na mesa" in rotulo:
            c.font = Font(bold=True, color="006100")
        linha += 1

    # Distribuicao por faixa
    linha += 1
    ws.cell(row=linha, column=1, value="Distribuicao por faixa").font = Font(bold=True, size=12, color=_AZUL)
    linha += 1
    for col, titulo in enumerate(["Faixa (Tema 1.234)", "Qtd processos"], start=1):
        _hdr(ws.cell(row=linha, column=col, value=titulo))
    linha += 1
    for faixa, label in _FAIXA_LABEL.items():
        qtd = metricas["distribuicao_faixa"].get(faixa, 0)
        ws.cell(row=linha, column=1, value=label).border = _BORDA
        cq = ws.cell(row=linha, column=2, value=qtd)
        cq.alignment = Alignment(horizontal="center")
        cq.border = _BORDA
        linha += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 26

    # --- Aba Processos ---
    wp = wb.create_sheet("Processos")
    wp.sheet_view.showGridLines = False
    cols = [
        ("Numero do processo", 26),
        ("Tribunal", 10),
        ("Comarca", 22),
        ("Classe", 26),
        ("CID", 8),
        ("Oncologico", 11),
        ("Custo anual estimado", 20),
        ("Faixa", 24),
        ("% Uniao", 10),
        ("Valor ressarcivel", 20),
        ("Justica", 12),
        ("Ajuizamento", 14),
    ]
    for col, (titulo, largura) in enumerate(cols, start=1):
        _hdr(wp.cell(row=1, column=col, value=titulo))
        wp.column_dimensions[get_column_letter(col)].width = largura

    for i, p in enumerate(metricas["processos"], start=2):
        valores = [
            p.numero_processo,
            (p.tribunal or "").upper(),
            p.comarca or "-",
            p.classe or "-",
            p.cid or "-",
            "Sim" if p.oncologico else "Nao",
            _brl(p.custo_anual_estimado),
            _FAIXA_LABEL.get(p.faixa, p.faixa or "-"),
            f"{(p.percentual_ressarcivel or 0):.0%}",
            _brl(p.valor_ressarcivel_estimado),
            p.justica_competente or "-",
            (p.data_ajuizamento or "")[:10],
        ]
        for col, v in enumerate(valores, start=1):
            c = wp.cell(row=i, column=col, value=v)
            c.border = _BORDA
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F2F2F2")

    wp.freeze_panes = "A2"
    wb.save(destino)
    return destino
